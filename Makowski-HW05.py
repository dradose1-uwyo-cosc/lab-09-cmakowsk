#Name Colter Makowski
# Lab Section: 15
# HW Number 5
import openpyxl
wb = openpyxl.Workbook()
from openpyxl.styles import PatternFill
sheet = wb.active
b = "1E3A5F"
r = "D32F2F"
br = "6E4B3A"
t = "F1C27D"
y = "FFEB3B"
fill_blue = PatternFill(patternType='solid', fgColor='1E3A5F')
fill_red = PatternFill(patternType='solid', fgColor='D32F2F')
fill_brown = PatternFill(patternType='solid', fgColor='6E4B3A')
fill_tan = PatternFill(patternType='solid', fgColor='F1C27D')
fill_yellow = PatternFill(patternType='solid', fgColor='FFEB3B')
for row in range(1, 19):
    sheet.row_dimensions[row].height = 50
for column in range(1, 15):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 10

cblue = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 
         'A13', 'A14', 'A15', 'A16', 'A17', 'A18', 'B1', 'B2', 'B3', 'B4', 'B5', 'B6', 
         'B7', 'B8', 'B9', 'B10', 'B15', 'B16', 'B18', 'C1', 'C2', 'C3', 'C4', 'C8', 
         'C9', 'C15', 'C18', 'D1', 'D2', 'D18', 'E1', 'E18', 'F1', 'F16', 'F17', 'F18', 
         'G1', 'G15', 'G16', 'G17', 'G18', 'H1', 'H15', 'H16', 'H17', 'H18', 'I1', 
         'I16', 'I17', 'I18', 'J1', 'J2', 'J9', 'J18', 'K1', 'K2', 'K4', 'K9', 'K18', 
         'L1', 'L2', 'L4', 'L8', 'L9', 'L15', 'L18', 'M1', 'M2', 'M3', 'M4', 'M5', 
         'M7', 'M8', 'M9', 'M10', 'M15', 'M16', 'M18', 'N1', 'N2', 'N3', 'N4', 'N5', 
         'N6', 'N7', 'N8', 'N9', 'N10', 'N11', 'N12', 'N13', 'N14', 'N15', 'N16', 
         'N17', 'N18']
for cells in cblue:
    sheet[cells].fill = fill_blue
cred = ['D3', 'D14', 'D15', 'E2', 'E3', 'E11', 'E12', 'E13', 'E14', 'E15', 'F2', 'F3',
        'F9', 'F10', 'F12', 'F13', 'F14', 'F15', 'G2', 'G3', 'G10', 'G11', 'G12', 'G13',
        'G14', 'H2', 'H3', 'H10', 'H11', 'H12', 'H13', 'H14', 'I2', 'I3', 'I9', 'I10',
        'I12', 'I13', 'I14', 'I15', 'J3', 'J11', 'J12', 'J13', 'J14', 'J15', 'K3', 'K14',
        'K15', 'L3']
for cells in cred:
    sheet[cells].fill = fill_red
cyellow = ['F11', 'I11']
for cells in cyellow:
    sheet[cells].fill = fill_yellow
ctan = ['B12', 'B13', 'B14', 'C12', 'C13', 'C14', 'D5', 'D6', 'D8', 'D13', 'E7', 'E8', 
        'F5', 'F7', 'F8', 'G4', 'G5', 'G6', 'G7', 'G8', 'H4', 'H5', 'H6', 'H7', 'H8', 
        'I6', 'I8', 'J4', 'J5', 'J8', 'K5', 'K6', 'K8', 'L5', 'L6', 'L12', 'L13', 
        'L14', 'M6', 'M12', 'M13', 'M14']
for cells in ctan:
    sheet[cells].fill = fill_tan
cbrown = ['B11', 'B17', 'C5', 'C6', 'C7', 'C10', 'C11', 'C16', 'C17', 'D4', 'D7', 'D9', 'D10', 'D11', 'D12', 'D16', 'D17', 'E4', 'E5', 'E6', 'E9', 'E10', 'E16', 'E17', 'F4', 'F6', 'G9', 'H9', 'I4', 'I5', 'I7', 'J6', 'J7', 'J10', 'J16', 'J17', 'K7', 'K10', 'K11', 'K12', 'K16', 'K17', 'L7', 'L10', 'L11', 'L16', 'L17', 'M11', 'M17']
for cells in cbrown:
    sheet[cells].fill = fill_brown
    

wb.save("PixelArt.xlsx")