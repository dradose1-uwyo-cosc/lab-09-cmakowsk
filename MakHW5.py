import openpyxl
wb = openpyxl.Workbook()
from openpyxl.styles import PatternFill
sheet = wb.active
b = "1E3A5F"
r = "D32F2F"
br = "6E4B3A"
t = "F1C27D"
y = "FFEB3B"
for row in range(1, 18):
    sheet.row_dimensions[row].height = 50
for column in range(1, 15):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 10

pixel_art = [
[b, b, b, b, b, b, b, b, b, b, b, b, b, b],
[b, b, b, b, r, r, r, r, r, b, b, b, b, b],
[b, b, b, r, r, r, r, r, r, r, r, r, b, b],
[b, b, b, br, br, br, t, t, br, t, b, b, b, b],
[b, b, br, t, br, t, t, t, br, t, t, t, b, b],
[b, b, br, t, br, br, t, t, t, br, t, t, t, b],
[b, b, br, br, t, t, t, t, br, br, br, br, b, b],
[b, b, b, t, t, t, t, t, t, t, t, b, b, b],
[b, b, b, br, br, r, br, br, r, b, b, b, b, b],
[b, b, br, br, br, r, r, r, r, br, br, br, b, b],
[b, br, br, br, r, y, r, r, y, r, br, br, br, b],
[b, t, t, br, r, r, r, r, r, r, br, t, t, b],
[b, t, t, t, r, r, r, r, r, r, t, t, t, b],
[b, t, t, r, r, r, r, r, r, r, r, t, t, b],
[b, b, b, r, r, r, b, b, r, r, r, b, b, b],
[b, b, br, br, br, b, b, b, b, br, br, br, b, b],
[b, br, br, br, br, b, b, b, b, br, br, br, br, b],
[b, b, b, b, b, b, b, b, b, b, b, b, b, b],
]
for row_index, row in enumerate(pixel_art, start=1):
    for col_index, color in enumerate(row, start=1):
        cell = sheet.cell(row=row_index, column=col_index)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    

wb.save("PixelArt.xlsx")